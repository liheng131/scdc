"""
Excel 表格解析器

使用 openpyxl 库解析 .xlsx/.xls 文件，按 Sheet 分块输出。

为什么使用 data_only=True：
- Excel 文件可能包含公式（如 =SUM(A1:A10)），默认读取公式文本而非计算结果
- data_only=True 读取公式的缓存计算结果，适合分析场景

为什么按 Sheet 分块：
- 多 Sheet 工作簿通常每个 Sheet 代表不同主题/数据集
- 分块后每个 Sheet 独立送入 CleanerAgent，互不干扰
"""

import io
from typing import BinaryIO
import openpyxl
from app.parsers.base import BaseParser
from app.schemas.parser import ParseResult, Chunk
from app.core.exceptions import BusinessException

class ExcelParser(BaseParser):
    async def parse(self, file_stream: BinaryIO, filename: str) -> ParseResult:
        try:
            wb = openpyxl.load_workbook(file_stream, data_only=True)
            full_text = []
            chunks = []
            chunk_idx = 1

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = [f"--- Sheet: {sheet_name} ---"]

                for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                    row_vals = [str(cell) for cell in row if cell is not None]
                    if row_vals:
                        row_str = " | ".join(row_vals)
                        sheet_text.append(row_str)

                if len(sheet_text) > 1:
                    content = "\n".join(sheet_text)
                    full_text.append(content)
                    chunks.append(Chunk(
                        index=chunk_idx,
                        content=content,
                        metadata={"sheet": sheet_name, "rows": len(sheet_text) - 1}
                    ))
                    chunk_idx += 1

            content = "\n\n".join(full_text)
            return ParseResult(
                filename=filename,
                file_type="xlsx",
                content=content,
                chunks=chunks,
                metadata={"sheet_names": wb.sheetnames}
            )
        except Exception as e:
            raise BusinessException(code=422, message=f"Failed to parse Excel file: {str(e)}")
